
"""
Betaalbestanden Generator
"""

import streamlit as st
import pandas as pd
import os
import tempfile
import datetime as dt
import time
from openpyxl import load_workbook
import re
from io import BytesIO
import random
import zipfile

# Pagina configuratie
st.set_page_config(
    page_title="Betaalbestanden Generator - Energiemissie",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
    :root {
        --primary-color: #667eea;
        --secondary-color: #764ba2;
        --success-color: #4ade80;
        --error-color: #ef4444;
    }

    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 1rem;
        color: white;
        margin-bottom: 2rem;
        text-align: center;
    }

    .main-header h1 {
        color: white !important;
        font-size: 3rem;
        margin: 0;
        font-weight: 800;
    }

    .main-header p {
        color: rgba(255, 255, 255, 0.9);
        font-size: 1.2rem;
        margin-top: 0.5rem;
    }

    .success-box {
        background: #dcfce7;
        border-left: 4px solid #16a34a;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }

    .error-box {
        background: #fee2e2;
        border-left: 4px solid #dc2626;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }

    .warning-box {
        background: #fef3c7;
        border-left: 4px solid #f59e0b;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }

    .info-box {
        background: #dbeafe;
        border-left: 4px solid #3b82f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 2rem;
    }

    .stTabs [data-baseweb="tab"] {
        font-weight: 600;
        font-size: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# LEVERANCIERS
SUPPLIERS = {
    "kenter": {
        "naam": "Kenter",
        "tabnaam": "NL59INGB0006779355_D",
        "tabnaam_credit": "NL59INGB0006779355_D",  # Zelfde tab, gefilterd op kolom N
        "color": "#FF6B6B"
    },
    "liander": {
        "naam": "Liander",
        "tabnaam": "NL95INGB0000005585_D",
        "tabnaam_credit": "NL95INGB0000005585_C",
        "color": "#4ECDC4"
    },
    "vattenfall": {
        "naam": "Vattenfall",
        "tabnaam": "NL42INGB0000827935_D",
        "tabnaam_credit": "NL14ABNA0242263240_C",
        "color": "#FFE66D"
    },
    "eneco": {
        "naam": "Eneco",
        "tabnaam": "NL13ABNA0640000797_D",
        "tabnaam_credit": "NL13ABNA0640000797_C",
        "color": "#95E1D3"
    },
    "vitens": {
        "naam": "Vitens",
        "tabnaam": "NL94INGB0000869000_D",
        "tabnaam_credit": "NL94INGB0000869000_C",
        "color": "#A8E6CF"
    }
}

# KLANT -> leveranciers mapping (afgedwongen)
CUSTOMER_SUPPLIERS = {
    "Provincie Noord-Holland": ["vattenfall", "kenter", "liander"],
    "GGZ Centraal": ["eneco", "vitens"]
}

# =====================================================================
# FUNCTIE OM RANDOM FACTUUR TE KRIJGEN
# =====================================================================
def get_random_invoice(bronbestand_bytes, tabnaam):
    """Haal een random factuurnummer op uit het bronbestand voor referentie"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(bronbestand_bytes)
            tmp_path = tmp.name

        try:
            df_raw = pd.read_excel(tmp_path, sheet_name=tabnaam, header=None).dropna(how="all")
            df_headers = df_raw.iloc[0]
            df_data = df_raw[1:].copy()
            df_data.columns = df_headers

            if len(df_data) > 0:
                factuurnummers = df_data.iloc[:, 2].astype(str).tolist()
                if factuurnummers:
                    return random.choice(factuurnummers)
            return None
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
    except Exception:
        return None

CREDIT_TYPES = ["credit", "correctie"]  # Deze gaan naar credit Betaalbestanden
DEBET_TYPE = "debet"  # Deze gaan naar debet Betaalbestanden

# Template opslag directory
TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "saved_templates")
os.makedirs(TEMPLATE_DIR, exist_ok=True)

def save_template_to_disk(customer_name, supplier_key, template_bytes, filename, is_credit=False):
    """Sla een template op naar disk voor permanent gebruik - NU KLANT-SPECIFIEK"""
    template_type = "credit" if is_credit else "debet"
    safe_customer = customer_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
    
    # Maak een veilige template filename
    safe_filename = filename.replace(" ", "_").replace("/", "_").replace("\\", "_")
    
    # BELANGRIJK: Gebruik klant-specifieke path
    template_path = os.path.join(TEMPLATE_DIR, f"{safe_customer}_{supplier_key}_{template_type}.xlsx")
    filename_path = os.path.join(TEMPLATE_DIR, f"{safe_customer}_{supplier_key}_{template_type}.txt")
    
    try:
        # Sla template bytes op
        with open(template_path, "wb") as f:
            f.write(template_bytes)
        
        # Sla originele filename op
        with open(filename_path, "w", encoding="utf-8") as f:
            f.write(filename)
        
        return True
    except Exception as e:
        st.error(f"Fout bij opslaan template: {e}")
        return False

def load_template_from_disk(customer_name, supplier_key, is_credit=False):
    """Laad een opgeslagen template van disk - NU KLANT-SPECIFIEK"""
    template_type = "credit" if is_credit else "debet"
    safe_customer = customer_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
    template_path = os.path.join(TEMPLATE_DIR, f"{safe_customer}_{supplier_key}_{template_type}.xlsx")
    filename_path = os.path.join(TEMPLATE_DIR, f"{safe_customer}_{supplier_key}_{template_type}.txt")
    
    try:
        if os.path.exists(template_path) and os.path.exists(filename_path):
            with open(template_path, "rb") as f:
                template_bytes = f.read()
            
            with open(filename_path, "r", encoding="utf-8") as f:
                filename = f.read()
            
            return template_bytes, filename
    except Exception as e:
        st.error(f"Fout bij laden template: {e}")
    
    return None, None

def delete_template_from_disk(customer_name, supplier_key, is_credit=False):
    """Verwijder een opgeslagen template van disk - NU KLANT-SPECIFIEK"""
    template_type = "credit" if is_credit else "debet"
    safe_customer = customer_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
    template_path = os.path.join(TEMPLATE_DIR, f"{safe_customer}_{supplier_key}_{template_type}.xlsx")
    filename_path = os.path.join(TEMPLATE_DIR, f"{safe_customer}_{supplier_key}_{template_type}.txt")
    
    try:
        deleted = False
        if os.path.exists(template_path):
            os.remove(template_path)
            deleted = True
        if os.path.exists(filename_path):
            os.remove(filename_path)
            deleted = True
        return deleted
    except Exception as e:
        st.error(f"Fout bij verwijderen template: {e}")
        return False

def list_saved_templates(customer_name=None):
    """Haal een lijst op van alle opgeslagen templates (optioneel gefilterd op klant)"""
    templates = []
    try:
        if not os.path.exists(TEMPLATE_DIR):
            return templates
        
        if customer_name:
            safe_customer = customer_name.replace(" ", "_").replace("/", "_").replace("\\", "_")
            prefix = f"{safe_customer}_"
        else:
            prefix = ""
            
        for file in os.listdir(TEMPLATE_DIR):
            if file.endswith(".txt"):
                if not prefix or file.startswith(prefix):
                    try:
                        with open(os.path.join(TEMPLATE_DIR, file), "r", encoding="utf-8") as f:
                            filename = f.read()
                            templates.append(filename)
                    except Exception:
                        continue
    except Exception as e:
        st.error(f"Fout bij laden template lijst: {e}")
    return templates

def has_credit_or_correctie_rows(bronbestand_bytes, tabnaam):
    """
    Controleer of een tabblad regels bevat met Credit of Correctie in kolom N
    """
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(bronbestand_bytes)
            tmp_path = tmp.name

        try:
            df_raw = pd.read_excel(tmp_path, sheet_name=tabnaam, header=None).dropna(how="all")
            df_headers = df_raw.iloc[0]
            df_data = df_raw[1:].copy()
            df_data.columns = df_headers
            
            # Check kolom N (14e kolom, 0-indexed = 13)
            if len(df_data.columns) < 14:
                return False
            
            kolom_n = df_data.iloc[:, 13]  # Kolom N is index 13 (A=0, B=1, ..., N=13)
            
            # Check of er credit/correctie waarden zijn
            mask = kolom_n.astype(str).str.lower().str.contains('credit|correctie', na=False, regex=True)
            return mask.any()
            
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
    except Exception:
        return False

def split_credit_correctie(df):
    """
    Split dataframe in debet en credit/correctie regels.
    Zoekt naar "Credit" of "Correctie" in kolom N (14e kolom)
    
    Kolom N waarden:
    - "Debet" → gaat naar DEBET Betaalbestanden
    - "Credit" → gaat naar CREDIT Betaalbestanden
    - "Correctie" → gaat naar CREDIT Betaalbestanden
    """
    # Check of er minstens 14 kolommen zijn (kolom N = index 13)
    if len(df.columns) < 14:
        # Geen kolom N, return alles als debet
        return df, pd.DataFrame(columns=df.columns)
    
    # Kolom N is de 14e kolom (index 13: A=0, B=1, ..., N=13)
    kolom_n = df.iloc[:, 13]
    
    # Check welke regels credit/correctie zijn (case-insensitive)
    # Alles wat "credit" of "correctie" bevat gaat naar credit Betaalbestanden
    mask = kolom_n.astype(str).str.lower().str.contains('credit|correctie', na=False, regex=True)
    return df[~mask].copy(), df[mask].copy()


def process_supplier(
    bronbestand_bytes,
    template_bytes,
    supplier_key: str,
    periode: str,
    credit: bool,
    template_filename: str = None,
    is_credit_sheet: bool = False,
    credit_template_bytes=None,
    credit_template_filename: str = None
):
    """Verwerk een enkele leverancier (debet of credit)"""

    if is_credit_sheet:
        tabnaam = SUPPLIERS[supplier_key]["tabnaam_credit"]
        template_to_use = credit_template_bytes if credit_template_bytes else template_bytes
        filename_to_use = credit_template_filename if credit_template_filename else template_filename
    else:
        tabnaam = SUPPLIERS[supplier_key]["tabnaam"]
        template_to_use = template_bytes
        filename_to_use = template_filename

    datum_excel = dt.date.today().strftime("%d%m%y")
    datum_tekst = dt.date.today().strftime("%d-%m-%Y")

    bron_path = None
    tpl_path = None

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as bron_tmp:
            bron_tmp.write(bronbestand_bytes)
            bron_path = bron_tmp.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tpl_tmp:
            tpl_tmp.write(template_to_use)
            tpl_path = tpl_tmp.name

        xls_file = pd.ExcelFile(bron_path)
        sheets = xls_file.sheet_names
        xls_file.close()

        if tabnaam not in sheets:
            try:
                if bron_path and os.path.exists(bron_path):
                    os.unlink(bron_path)
                if tpl_path and os.path.exists(tpl_path):
                    os.unlink(tpl_path)
            except Exception:
                pass

            return {
                "success": False,
                "message": f"Tabblad {tabnaam} niet gevonden",
                "supplier": SUPPLIERS[supplier_key]["naam"],
                "is_credit": is_credit_sheet
            }

        df_raw = pd.read_excel(bron_path, sheet_name=tabnaam, header=None).dropna(how="all")
        df_headers = df_raw.iloc[0]
        df_data_all = df_raw[1:].copy()
        df_data_all.columns = df_headers

        # Split credit/correctie van debet regels
        df_debet, df_credit = split_credit_correctie(df_data_all)

        # Kies de juiste data op basis van is_credit_sheet
        if is_credit_sheet:
            df_data = df_credit
            type_label = "credit/correctie"
        else:
            df_data = df_debet
            type_label = "debet"

        if len(df_data) == 0:
            try:
                if bron_path and os.path.exists(bron_path):
                    os.unlink(bron_path)
                if tpl_path and os.path.exists(tpl_path):
                    os.unlink(tpl_path)
            except Exception:
                pass

            return {
                "success": False,
                "message": f"Geen {type_label} data gevonden in tabblad",
                "supplier": SUPPLIERS[supplier_key]["naam"],
                "is_credit": is_credit_sheet,
                "no_data": True
            }

        wb = load_workbook(tpl_path)
        ws_spec = wb["Specificatie"]

        # Verwijder oude data (behoud header)
        last = 1
        while ws_spec.cell(last + 1, 1).value not in (None, ""):
            last += 1
        if last > 1:
            ws_spec.delete_rows(2, last - 1)

        # FIX: Kopieer data in volgorde - kopieer alle kolommen zoals ze zijn
        for row_idx, row_data in enumerate(df_data.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws_spec.cell(row_idx, col_idx, value)

        ws_verz = wb["Betaalbestanden"]
        ws_verz["B4"].value = datum_tekst

        if supplier_key in ("kenter", "liander", "vattenfall"):
            ws_verz["C24"].value = f"{supplier_key.capitalize()}_{datum_excel}"
        else:
            old = ws_verz["C24"].value
            if isinstance(old, str):
                new = re.sub(r"\d{6,8}", datum_excel, old, 1)
                ws_verz["C24"].value = new
            else:
                ws_verz["C24"].value = f"{supplier_key.upper()}_{datum_excel}"

        ws_verz["C26"].value = periode

        if filename_to_use:
            template_basename = filename_to_use.replace(".xlsx", "")
        else:
            template_basename = SUPPLIERS[supplier_key]["naam"]
            if is_credit_sheet:
                template_basename += "_Credit"

        pattern = r"_(\d{6,8})(?!.*_\d)"

        if re.search(pattern, template_basename):
            output_filename_base = re.sub(pattern, f"_{datum_excel}", template_basename)
        else:
            output_filename_base = f"{template_basename}_{datum_excel}"

        output_filename = f"{output_filename_base}.xlsx"

        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)
        excel_bytes = output.read()

        wb_check = load_workbook(BytesIO(excel_bytes))
        ws_new = wb_check["Specificatie"]

        rows_new = []
        r = 2
        while ws_new.cell(r, 1).value not in (None, ""):
            row = [ws_new.cell(r, c).value for c in range(1, df_data.shape[1] + 1)]
            rows_new.append(row)
            r += 1

        wb_check.close()
        df_spec_new = pd.DataFrame(rows_new, columns=df_data.columns)

        kol_excl = df_data.columns.get_loc("Excl. BTW")
        kol_btw = df_data.columns.get_loc("BTW")
        kol_incl = df_data.columns.get_loc("Incl. BTW")

        sum_excl_bron = df_data.iloc[:, kol_excl].astype(float).sum()
        sum_btw_bron = df_data.iloc[:, kol_btw].astype(float).sum()
        sum_incl_bron = df_data.iloc[:, kol_incl].astype(float).sum()

        sum_excl_new = df_spec_new.iloc[:, kol_excl].astype(float).sum()
        sum_btw_new = df_spec_new.iloc[:, kol_btw].astype(float).sum()
        sum_incl_new = df_spec_new.iloc[:, kol_incl].astype(float).sum()

        bedragen_kloppen = (
            abs(sum_excl_bron - sum_excl_new) < 0.01 and
            abs(sum_btw_bron - sum_btw_new) < 0.01 and
            abs(sum_incl_bron - sum_incl_new) < 0.01
        )

        try:
            if bron_path and os.path.exists(bron_path):
                os.unlink(bron_path)
        except Exception:
            pass

        try:
            if tpl_path and os.path.exists(tpl_path):
                os.unlink(tpl_path)
        except Exception:
            pass

        pdf_bytes = None
        pdf_filename = None

        if bedragen_kloppen:
            try:
                import win32com.client as win32
                import pythoncom
                
                # Initialiseer COM voor deze thread
                pythoncom.CoInitialize()
                
                print(f"Starting PDF generation for {SUPPLIERS[supplier_key]['naam']}...")

                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
                    tmp_excel.write(excel_bytes)
                    tmp_excel_path = tmp_excel.name
                print(f"Created temp Excel file: {tmp_excel_path}")

                pdf_temp_path = tmp_excel_path.replace(".xlsx", ".pdf")
                print(f"PDF will be saved to: {pdf_temp_path}")

                try:
                    print("Dispatching Excel...")
                    excel_app = win32.Dispatch("Excel.Application")
                    excel_app.Visible = False
                    excel_app.DisplayAlerts = False
                    print("Excel dispatched successfully")

                    print(f"Opening workbook: {tmp_excel_path}")
                    wb_pdf = excel_app.Workbooks.Open(tmp_excel_path, ReadOnly=True)
                    print("Workbook opened")
                    
                    print("Getting Betaalbestanden worksheet...")
                    ws_pdf = wb_pdf.Worksheets("Betaalbestanden")
                    print("Worksheet found")
                    
                    print(f"Exporting to PDF: {pdf_temp_path}")
                    ws_pdf.ExportAsFixedFormat(0, pdf_temp_path)
                    print("PDF exported successfully!")

                    wb_pdf.Close(False)
                    excel_app.Quit()
                    print("Excel closed")

                    with open(pdf_temp_path, "rb") as f:
                        pdf_bytes = f.read()
                    print(f"PDF read successfully, size: {len(pdf_bytes)} bytes")

                    pdf_filename = output_filename.replace(".xlsx", ".pdf")

                    try:
                        os.unlink(tmp_excel_path)
                        os.unlink(pdf_temp_path)
                        print("Temp files cleaned up")
                    except Exception as cleanup_error:
                        print(f"Cleanup warning: {cleanup_error}")

                except Exception as pdf_error:
                    # Log de fout maar laat de verwerking doorgaan
                    print(f"PDF generatie fout voor {SUPPLIERS[supplier_key]['naam']}: {pdf_error}")
                    import traceback
                    print(traceback.format_exc())
                    try:
                        if "wb_pdf" in locals():
                            wb_pdf.Close(False)
                        if "excel_app" in locals():
                            excel_app.Quit()
                        if os.path.exists(tmp_excel_path):
                            os.unlink(tmp_excel_path)
                        if os.path.exists(pdf_temp_path):
                            os.unlink(pdf_temp_path)
                    except Exception:
                        pass
                finally:
                    # Uninitialize COM
                    try:
                        pythoncom.CoUninitialize()
                    except Exception:
                        pass

            except ImportError as import_error:
                # Win32com niet beschikbaar - PDF generatie niet mogelijk
                print(f"PDF generatie niet beschikbaar: pywin32 niet geïnstalleerd")
            except Exception as e:
                print(f"PDF generatie fout: {e}")
                import traceback
                print(traceback.format_exc())

        return {
            "success": bedragen_kloppen,
            "message": "Bedragen kloppen ✓" if bedragen_kloppen else "Bedragen kloppen NIET ✗",
            "supplier": SUPPLIERS[supplier_key]["naam"],
            "excl": sum_excl_new,
            "btw": sum_btw_new,
            "incl": sum_incl_new,
            "excl_bron": sum_excl_bron,
            "btw_bron": sum_btw_bron,
            "incl_bron": sum_incl_bron,
            "excel_bytes": excel_bytes,
            "filename": output_filename,
            "pdf_bytes": pdf_bytes,
            "pdf_filename": pdf_filename,
            "is_credit": is_credit_sheet
        }

    except Exception as e:
        try:
            if bron_path and os.path.exists(bron_path):
                os.unlink(bron_path)
        except Exception:
            pass

        try:
            if tpl_path and os.path.exists(tpl_path):
                os.unlink(tpl_path)
        except Exception:
            pass

        return {
            "success": False,
            "message": f"Fout: {str(e)}",
            "supplier": SUPPLIERS[supplier_key]["naam"],
            "is_credit": is_credit_sheet
        }

# =====================================================
# Euromaster SEFE verwerking
# =====================================================
def init_validation_report():
    return {"correcties": [], "waarschuwingen": [], "fouten": []}

def euromaster_read_csv(csv_bytes, report):
    try:
        df = pd.read_csv(
            BytesIO(csv_bytes),
            sep=";",
            dtype=str
        )

        # kolomnamen opschonen
        df.columns = (
            df.columns
            .str.strip()
            .str.replace("\ufeff", "", regex=False)
        )

        report["correcties"].append("Csv bestand geladen")
        return df

    except Exception as e:
        report["fouten"].append(f"Fout bij lezen csv bestand: {e}")
        return None

def euromaster_check_and_fix_amounts(df, report):
    gewijzigd = False

    required_cols = ["AmountInDocCurr", "Debit/Credit", "TaxAmount"]
    for col in required_cols:
        if col not in df.columns:
            report["fouten"].append(f"Kolom {col} niet gevonden")
            return df, False

    # bedragen opschonen
    for col in ["AmountInDocCurr", "TaxAmount"]:
        df[col] = (
            df[col]
            .astype(str)
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
            .str.strip()
        )
        df[col] = pd.to_numeric(df[col], errors="coerce")

        if df[col].isna().any():
            report["fouten"].append(f"Kolom {col} bevat ongeldige waarden")
            return df, False

    boeking = df[df["Debit/Credit"] == "C"]["AmountInDocCurr"].sum()

    tegen_excl = df[df["Debit/Credit"] == "D"]["AmountInDocCurr"].sum()
    tegen_btw = df[df["Debit/Credit"] == "D"]["TaxAmount"].sum()

    tegen_totaal = round(tegen_excl + tegen_btw, 2)
    verschil = round(boeking - tegen_totaal, 2)

    if abs(verschil) > 0.01:
        report["fouten"].append(
            f"Boeking ({boeking}) en tegenboekingen ({tegen_totaal}) komen niet overeen"
        )
        return df, False

    if abs(verschil) > 0:
        idx = df[df["Debit/Credit"] == "D"].index[0]
        df.loc[idx, "AmountInDocCurr"] += verschil
        gewijzigd = True
        report["correcties"].append(
            f"Centverschil {verschil} gealloceerd in tegenboekingsregel"
        )

    return df, gewijzigd



def euromaster_fix_kostenplaatsen(df, report):
    """
    Verwacht kolom: Kostenplaats
    Meerdere kostenplaatsen gescheiden door komma
    """
    gewijzigd = False
    rows = []

    for _, row in df.iterrows():
        if str(row.get("Type", "")) != "Tegenboeking":
            rows.append(row)
            continue

        kp_raw = str(row.get("Kostenplaats", "")).strip()
        if kp_raw == "" or kp_raw.lower() == "nan":
            report["fouten"].append("Tegenboeking zonder kostenplaats gevonden. Contact opnemen met Euromaster")
            continue

        kps = [k.strip() for k in kp_raw.split(",") if k.strip()]
        if len(kps) == 1:
            rows.append(row)
        else:
            try:
                bedrag = round(float(row["N"]) / len(kps), 2)
            except Exception:
                report["fouten"].append("Kon bedrag niet verdelen over kostenplaatsen")
                continue

            for kp in kps:
                new_row = row.copy()
                new_row["Kostenplaats"] = kp
                new_row["N"] = bedrag
                rows.append(new_row)

            gewijzigd = True
            report["correcties"].append(f"Tegenboeking opgesplitst in {len(kps)} regels")

    return pd.DataFrame(rows), gewijzigd

def euromaster_fix_factuurnummer_column_f(df, report):
    """
    Stap 5:
    Alleen bij wijzigingen: kolom F converteren naar getalnotitie zonder decimalen
    """
    try:
        df["F"] = (
            df["F"]
            .astype(str)
            .str.replace(".0", "", regex=False)
            .astype(int)
        )
        report["correcties"].append("Factuurnummer kolom F geconverteerd naar getal zonder decimalen")
        return df, True
    except Exception as e:
        report["fouten"].append(f"Fout bij converteren factuurnummer kolom F: {e}")
        return df, False

def preprocess_euromaster_sefe(csv_bytes):
    report = init_validation_report()
    gewijzigd = False

    df = euromaster_read_csv(csv_bytes, report)
    if df is None:
        return None, report, False

    df, changed_amounts = euromaster_check_and_fix_amounts(df, report)
    gewijzigd = gewijzigd or changed_amounts

    df, changed_kp = euromaster_fix_kostenplaatsen(df, report)
    gewijzigd = gewijzigd or changed_kp

    if report["fouten"]:
        return None, report, False

    if gewijzigd:
        df, ok = euromaster_fix_factuurnummer_column_f(df, report)
        if not ok:
            return None, report, False

        out = BytesIO()
        df.to_excel(out, index=False)
        out.seek(0)
        return out.read(), report, True

    report["waarschuwingen"].append(
        "Geen wijzigingen nodig. Bestand kan zonder opslaan worden afgesloten."
    )
    return None, report, False


# =====================================================
# Helper: per klant state in session_state
# =====================================================
def get_customer_state(customer_name):
    if "customer_states" not in st.session_state:
        st.session_state.customer_states = {}

    if customer_name not in st.session_state.customer_states:
        st.session_state.customer_states[customer_name] = {
            "bronbestand": None,
            "templates": {},
            "template_filenames": {},
            "credit_templates": {},
            "credit_template_filenames": {},
            "supplier_settings": {key: {"selected": False, "credit": False, "periode": ""} for key in SUPPLIERS.keys()},
            "results": []
        }

    return st.session_state.customer_states[customer_name]

# =====================================================
# UI: standaard flow voor Provincie Noord-Holland en GGZ Centraal
# =====================================================
def render_standard_customer_flow(customer_name, allowed_supplier_keys):
    state = get_customer_state(customer_name)

    # Sidebar met info (per klant)
    with st.sidebar:
        st.markdown(f"## {customer_name}")
        st.markdown("---")
        
        # Reset knop bovenaan
        st.markdown("### 🔄 Nieuw Betaalbestanden")
        if st.button("➕ Maak nieuw Betaalbestanden", type="primary", use_container_width=True, key=f"reset_{customer_name}"):
            # Reset alle state voor deze klant
            if "customer_states" in st.session_state and customer_name in st.session_state.customer_states:
                # Behoud alleen de templates
                templates = st.session_state.customer_states[customer_name].get("templates", {})
                template_filenames = st.session_state.customer_states[customer_name].get("template_filenames", {})
                credit_templates = st.session_state.customer_states[customer_name].get("credit_templates", {})
                credit_template_filenames = st.session_state.customer_states[customer_name].get("credit_template_filenames", {})
                
                # Reset alles behalve templates
                st.session_state.customer_states[customer_name] = {
                    "bronbestand": None,
                    "templates": templates,
                    "template_filenames": template_filenames,
                    "credit_templates": credit_templates,
                    "credit_template_filenames": credit_template_filenames,
                    "supplier_settings": {key: {"selected": False, "credit": False, "periode": ""} for key in SUPPLIERS.keys()},
                    "results": []
                }
                
                # Verwijder ook de file uploader state zodat de UI reset
                upload_key = f"{customer_name}_bron_upload"
                if upload_key in st.session_state:
                    del st.session_state[upload_key]
                
                st.success("✅ Nieuw Betaalbestanden gestart!")
                time.sleep(0.5)  # Korte delay voor de success message
                st.rerun()
        
        st.markdown("---")
        st.markdown("### 📋 Instructies")
        st.markdown("""
        1. **Tab 1**: Upload bronbestand
        2. **Tab 2**: Upload templates die nog niet eerder zijn gebruikt
        3. **Tab 3**: Configureer periode
        4. **Tab 4**: Start verwerking & Download resultaten
        """)
        st.markdown("---")
        st.markdown("### 💾 Opgeslagen Templates")
        
        # Haal lijst van opgeslagen templates op - GEFILTERD OP KLANT
        saved_templates = list_saved_templates(customer_name)
        
        if len(saved_templates) > 0:
            st.success(f"✅ {len(saved_templates)} template(s) permanent opgeslagen")
            
            # Toon lijst van opgeslagen templates
            with st.expander("📄 Bekijk opgeslagen templates"):
                for filename in saved_templates:
                    st.text(f"• {filename}")
        else:
            st.info("Nog geen templates opgeslagen")
        
        st.markdown("---")
        
        # Check of pywin32 beschikbaar is
        st.markdown("---")
        st.markdown("### 📄 PDF Generatie")
        try:
            import win32com.client
            st.success("✅ Excel beschikbaar - PDF's worden gegenereerd")
        except ImportError:
            st.warning("⚠️ PDF generatie niet beschikbaar")
            st.markdown("""
            **Installeer pywin32:**
            ```
            pip install pywin32
            ```
            Zie `PDF_INSTALLATIE.md` voor details.
            """)

    tab1, tab2, tab3, tab4 = st.tabs(["📁 Upload", "📋 Templates", "⚙️ Configuratie", "🚀 Verwerken & Resultaten"])

    # TAB 1: Upload bronbestand
    with tab1:
        st.markdown("## 📁 Bronbestand uploaden")
        st.markdown("Upload het exportbestand vanuit DB Energie (.xlsx)")

        uploaded_bron = st.file_uploader(
            "Kies bronbestand",
            type=["xlsx", "xls"],
            key=f"{customer_name}_bron_upload",
            help="Dit is het bestand dat je exporteert uit DB Energie"
        )

        if uploaded_bron:
            state["bronbestand"] = uploaded_bron.read()
            # Resultaten blijven behouden - alleen verwijderen via reset knop

            st.success(f"✅ Bestand geladen: **{uploaded_bron.name}**")
            
            # Waarschuw als er al resultaten zijn
            if state["results"]:
                st.warning("⚠️ Er zijn nog oude resultaten beschikbaar in Tab 4. Wil je opnieuw beginnen? Klik op '➕ Maak nieuw Betaalbestanden' in de sidebar.")

            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(state["bronbestand"])
                    tmp_path = tmp.name

                xls = pd.ExcelFile(tmp_path)
                sheets = set(xls.sheet_names)
                xls.close()

                st.markdown("### 🔍 Gevonden leveranciers (binnen deze klant):")

                # Reset selections eerst
                for key in SUPPLIERS.keys():
                    state["supplier_settings"][key]["selected"] = False
                    state["supplier_settings"][key]["credit"] = False

                # Debet
                st.markdown("**📊 Debet facturen:**")
                cols_debet = st.columns(len(allowed_supplier_keys))

                for idx, key in enumerate(allowed_supplier_keys):
                    supplier = SUPPLIERS[key]
                    with cols_debet[idx]:
                        if supplier["tabnaam"] in sheets:
                            st.markdown(f"✅ **{supplier['naam']}**")
                            state["supplier_settings"][key]["selected"] = True
                        else:
                            st.markdown(f"❌ {supplier['naam']}")
                            state["supplier_settings"][key]["selected"] = False

                # Credit
                st.markdown("**💳 Credit/Correctie facturen:**")
                cols_credit = st.columns(len(allowed_supplier_keys))

                has_credits = False
                for idx, key in enumerate(allowed_supplier_keys):
                    supplier = SUPPLIERS[key]
                    with cols_credit[idx]:
                        # Check 1: Is er een apart _C tabblad?
                        has_credit_tab = supplier["tabnaam_credit"] in sheets
                        
                        # Check 2: Zijn er Credit/Correctie regels in het _D tabblad?
                        has_credit_rows = False
                        if supplier["tabnaam"] in sheets:
                            has_credit_rows = has_credit_or_correctie_rows(state["bronbestand"], supplier["tabnaam"])
                        
                        if has_credit_tab or has_credit_rows:
                            if has_credit_tab and has_credit_rows:
                                st.markdown(f"✅ **{supplier['naam']}** (tab + regels)")
                            elif has_credit_tab:
                                st.markdown(f"✅ **{supplier['naam']}** (tab)")
                            else:
                                st.markdown(f"✅ **{supplier['naam']}** (regels)")
                            
                            state["supplier_settings"][key]["credit"] = True
                            has_credits = True
                        else:
                            st.markdown(f"⚪ {supplier['naam']}")

                if has_credits:
                    st.info("💡 Credit/Correctie facturen gevonden. Upload credit templates in de Configuratie tab.")

            except Exception as e:
                st.error(f"Fout bij het lezen van het bestand: {e}")
            finally:
                if tmp_path:
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass
        else:
            st.info("👆 Upload eerst een bronbestand om te beginnen")

    # TAB 2: Template Management
    with tab2:
        st.markdown("## 📋 Template Management")
        st.markdown("Upload hier alle templates. Deze worden permanent opgeslagen en automatisch gebruikt.")
        
        st.markdown("---")
        
        # Toon voor elke leverancier de template upload mogelijkheid
        for supplier_key in allowed_supplier_keys:
            supplier = SUPPLIERS[supplier_key]
            
            with st.expander(f"📁 {supplier['naam']} Templates", expanded=False):
                col1, col2 = st.columns(2)
                
                # DEBET TEMPLATE
                with col1:
                    st.markdown("### 📊 Debet Template")
                    
                    # Check of er al een opgeslagen template is
                    saved_template_bytes, saved_template_filename = load_template_from_disk(customer_name, supplier_key, is_credit=False)
                    
                    if saved_template_bytes and saved_template_filename:
                        st.success(f"✅ Opgeslagen: {saved_template_filename}")
                        
                        if st.button("🗑️ Verwijder debet template", key=f"tmpl_del_debet_{customer_name}_{supplier_key}"):
                            if delete_template_from_disk(customer_name, supplier_key, is_credit=False):
                                st.success("Template verwijderd!")
                                st.rerun()
                    else:
                        st.info("Nog geen debet template opgeslagen")
                    
                    # Upload nieuwe template
                    debet_upload = st.file_uploader(
                        "Upload debet template",
                        type=["xlsx"],
                        key=f"tmpl_upload_debet_{customer_name}_{supplier_key}"
                    )
                    
                    if debet_upload:
                        # Check of dit een nieuw bestand is (niet al verwerkt)
                        upload_key = f"processed_debet_{customer_name}_{supplier_key}_{debet_upload.name}"
                        
                        if upload_key not in st.session_state:
                            template_bytes = debet_upload.read()
                            if save_template_to_disk(customer_name, supplier_key, template_bytes, debet_upload.name, is_credit=False):
                                st.success(f"✅ Template opgeslagen: {debet_upload.name}")
                                st.session_state[upload_key] = True
                                time.sleep(1)
                                st.rerun()
                            else:
                                st.error("Kon template niet opslaan")
                
                # CREDIT TEMPLATE
                with col2:
                    st.markdown("### 💳 Credit Template")
                    
                    # Check of er al een opgeslagen credit template is
                    saved_credit_bytes, saved_credit_filename = load_template_from_disk(customer_name, supplier_key, is_credit=True)
                    
                    if saved_credit_bytes and saved_credit_filename:
                        st.success(f"✅ Opgeslagen: {saved_credit_filename}")
                        
                        if st.button("🗑️ Verwijder credit template", key=f"tmpl_del_credit_{customer_name}_{supplier_key}"):
                            if delete_template_from_disk(customer_name, supplier_key, is_credit=True):
                                st.success("Template verwijderd!")
                                st.rerun()
                    else:
                        st.info("Nog geen credit template opgeslagen")
                    
                    # Upload nieuwe credit template
                    credit_upload = st.file_uploader(
                        "Upload credit template",
                        type=["xlsx"],
                        key=f"tmpl_upload_credit_{customer_name}_{supplier_key}"
                    )
                    
                    if credit_upload:
                        # Check of dit een nieuw bestand is (niet al verwerkt)
                        upload_key = f"processed_credit_{customer_name}_{supplier_key}_{credit_upload.name}"
                        
                        if upload_key not in st.session_state:
                            credit_bytes = credit_upload.read()
                            if save_template_to_disk(customer_name, supplier_key, credit_bytes, credit_upload.name, is_credit=True):
                                st.success(f"✅ Template opgeslagen: {credit_upload.name}")
                                st.session_state[upload_key] = True
                                time.sleep(1)
                                st.rerun()
                            else:
                                st.error("Kon template niet opslaan")
        
        st.markdown("---")
        st.info("💡 Tip: Upload alle templates hier één keer. Ze worden permanent opgeslagen en automatisch gebruikt in de Configuratie tab!")

    # TAB 3: Configuratie
    with tab3:
        st.markdown("## ⚙️ Configureer leveranciers")
        st.info("💡 Templates worden automatisch geladen vanuit Tab 2 (Templates). Je hoeft ze hier niet meer te uploaden! Mocht je nieuwe templates willen toevoegen, doe dit in Tab 2. ")
        
        if not state["bronbestand"]:
            st.warning("⚠️ Upload eerst een bronbestand in de Upload tab")
        else:
            for supplier_key in allowed_supplier_keys:
                supplier = SUPPLIERS[supplier_key]
                settings = state["supplier_settings"][supplier_key]

                with st.expander(
                    f"🔌 {supplier['naam']}" + (" ✅ Actief" if settings["selected"] else " ⏸️ Inactief"),
                    expanded=settings["selected"]
                ):
                    col1, col2 = st.columns([1, 3])

                    with col1:
                        selected = st.checkbox(
                            "Verwerken",
                            value=settings["selected"],
                            key=f"{customer_name}_sel_{supplier_key}"
                        )
                        settings["selected"] = selected

                    if selected:
                        with col2:
                            has_credit = settings["credit"]
                            if has_credit:
                                st.success(f"💳 Credit/Correctie facturen gevonden")

                        st.markdown("---")
                        st.markdown("**🔍 Referentie factuurnummer:**")

                        random_invoice_debet = get_random_invoice(state["bronbestand"], supplier["tabnaam"])
                        if random_invoice_debet:
                            st.markdown(f"""
                            <div class="info-box">
                                <strong>📊 Debet:</strong> {random_invoice_debet}<br>
                                <em style="font-size: 0.9em; color: #6b7280;">Gebruik dit factuurnummer om de periode op te zoeken in DB EFactuur</em>
                            </div>
                            """, unsafe_allow_html=True)

                        if has_credit:
                            random_invoice_credit = get_random_invoice(state["bronbestand"], supplier["tabnaam_credit"])
                            if random_invoice_credit:
                                st.markdown(f"""
                                <div class="info-box">
                                    <strong>💳 Credit:</strong> {random_invoice_credit}<br>
                                    <em style="font-size: 0.9em; color: #6b7280;">Gebruik dit factuurnummer om de periode op te zoeken in DB EFactuur</em>
                                </div>
                                """, unsafe_allow_html=True)

                        st.markdown("---")

                        periode = st.text_input(
                            "Periode (bijv. 01-12-2025 t/m 31-12-2025)",
                            value=settings["periode"],
                            key=f"{customer_name}_periode_{supplier_key}",
                            help="Zoek het factuurnummer hierboven op in DB EFactuur om de juiste periode te vinden"
                        )
                        settings["periode"] = periode
                        
                        st.markdown("---")
                        st.markdown("**📋 Templates:**")
                        
                        # Laad automatisch templates van disk
                        saved_debet_bytes, saved_debet_filename = load_template_from_disk(customer_name, supplier_key, is_credit=False)
                        
                        if saved_debet_bytes and saved_debet_filename:
                            st.success(f"✅ Debet template: {saved_debet_filename}")
                            state["templates"][supplier_key] = saved_debet_bytes
                            state["template_filenames"][supplier_key] = saved_debet_filename
                        else:
                            st.warning(f"⚠️ Geen debet template gevonden. Upload in Tab 2 (Templates)!")
                        
                        if has_credit:
                            saved_credit_bytes, saved_credit_filename = load_template_from_disk(customer_name, supplier_key, is_credit=True)
                            
                            if saved_credit_bytes and saved_credit_filename:
                                st.success(f"✅ Credit template: {saved_credit_filename}")
                                state["credit_templates"][supplier_key] = saved_credit_bytes
                                state["credit_template_filenames"][supplier_key] = saved_credit_filename
                            else:
                                st.warning(f"⚠️ Geen credit template gevonden. Upload in Tab 2 (Templates)!")


    # TAB 4: Verwerken & Resultaten
    with tab4:
        st.markdown("## 🚀 Bestanden verwerken & Resultaten")
        
        # Toon resultaten als die er zijn
        if state["results"]:
            total_processed = len([r for r in state["results"] if not r.get("no_data")])
            successful_count = len([r for r in state["results"] if r["success"] and not r.get("no_data")])
            
            st.success(f"✅ {successful_count} van {total_processed} bestand(en) succesvol verwerkt")
            
            successful_results = [r for r in state["results"] if r["success"] and not r.get("no_data")]
            
            if len(successful_results) > 0:
                st.markdown("### 📥 Download alle bestanden")
                
                zip_buffer = BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for result in successful_results:
                        zip_file.writestr(result["filename"], result["excel_bytes"])
                        if result.get("pdf_bytes"):
                            zip_file.writestr(result["pdf_filename"], result["pdf_bytes"])
                
                zip_buffer.seek(0)
                
                st.download_button(
                    "📦 DOWNLOAD ALLE BESTANDEN (ZIP)",
                    data=zip_buffer.getvalue(),
                    file_name=f"{customer_name.replace(' ', '_')}_Betaalbestandenen_{dt.date.today().strftime('%Y%m%d')}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    key=f"{customer_name}_download_all_zip",
                    type="primary"
                )
                
                excel_count = len(successful_results)
                pdf_count = len([r for r in successful_results if r.get("pdf_bytes")])
                total_count = excel_count + pdf_count
                
                st.info(f"📦 De ZIP bevat {total_count} bestanden: {excel_count} Excel en {pdf_count} PDF")
                
                st.markdown("---")
                st.markdown("### 📋 Gedetailleerde resultaten")
                
                for result in state["results"]:
                    type_label = "💳 CREDIT" if result.get("is_credit") else "📊 DEBET"

                    if result.get("no_data"):
                        continue

                    if result["success"]:
                        st.markdown(f"""
                        <div class="success-box">
                            <h3 style="margin:0;">✅ {result['supplier']} {type_label}</h3>
                            <p style="margin:0.5rem 0 0 0;">{result['message']}</p>
                        </div>
                        """, unsafe_allow_html=True)

                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("Excl. BTW", f"€ {result['excl']:,.2f}")
                        with col2:
                            st.metric("BTW", f"€ {result['btw']:,.2f}")
                        with col3:
                            st.metric("Incl. BTW", f"€ {result['incl']:,.2f}")

                        st.markdown("**📥 Downloads:**")
                        col_excel, col_pdf = st.columns(2)

                        with col_excel:
                            st.download_button(
                                f"📊 {result['filename']}",
                                data=result["excel_bytes"],
                                file_name=result["filename"],
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"{customer_name}_excel_{result['supplier']}_{result.get('is_credit', False)}"
                            )

                        with col_pdf:
                            if result.get("pdf_bytes"):
                                st.download_button(
                                    f"📄 {result['pdf_filename']}",
                                    data=result["pdf_bytes"],
                                    file_name=result["pdf_filename"],
                                    mime="application/pdf",
                                    use_container_width=True,
                                    key=f"{customer_name}_pdf_{result['supplier']}_{result.get('is_credit', False)}"
                                )
                            else:
                                st.info("PDF: Excel niet beschikbaar", icon="ℹ️")

                    else:
                        st.markdown(f"""
                        <div class="error-box">
                            <h3 style="margin:0;">❌ {result['supplier']} {type_label}</h3>
                            <p style="margin:0.5rem 0 0 0;">{result['message']}</p>
                        </div>
                        """, unsafe_allow_html=True)

                        if "excl" in result:
                            st.markdown("**Vergelijking bedragen:**")

                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.markdown("**Bronbestand:**")
                                st.text(f"Excl: € {result['excl_bron']:,.2f}")
                                st.text(f"BTW:  € {result['btw_bron']:,.2f}")
                                st.text(f"Incl: € {result['incl_bron']:,.2f}")

                            with col2:
                                st.markdown("**Nieuw:**")
                                st.text(f"Excl: € {result['excl']:,.2f}")
                                st.text(f"BTW:  € {result['btw']:,.2f}")
                                st.text(f"Incl: € {result['incl']:,.2f}")

                            with col3:
                                st.markdown("**Verschil:**")
                                st.text(f"Excl: € {result['excl']-result['excl_bron']:,.2f}")
                                st.text(f"BTW:  € {result['btw']-result['btw_bron']:,.2f}")
                                st.text(f"Incl: € {result['incl']-result['incl_bron']:,.2f}")

                    st.markdown("---")
            
            st.markdown("---")
        
        # Verwerk sectie
        st.markdown("### ⚙️ Nieuwe verwerking starten")

        selected_suppliers = [
            key for key in allowed_supplier_keys
            if state["supplier_settings"][key]["selected"]
        ]

        missing_templates = [
            SUPPLIERS[key]["naam"] for key in selected_suppliers
            if key not in state["templates"]
        ]

        missing_credit_templates = [
            SUPPLIERS[key]["naam"] for key in selected_suppliers
            if state["supplier_settings"][key]["credit"] and key not in state["credit_templates"]
        ]

        can_process = (
            state["bronbestand"] and
            len(selected_suppliers) > 0 and
            len(missing_templates) == 0 and
            len(missing_credit_templates) == 0
        )

        if not can_process:
            st.warning("⚠️ Vul eerst alle vereiste velden in:")
            issues = []
            if not state["bronbestand"]:
                issues.append("Upload bronbestand")
            if len(selected_suppliers) == 0:
                issues.append("Selecteer minimaal 1 leverancier")
            if missing_templates:
                issues.append(f"Upload debet templates voor: {', '.join(missing_templates)}")
            if missing_credit_templates:
                issues.append(f"Upload credit templates voor: {', '.join(missing_credit_templates)}")

            for it in issues:
                st.write("•", it)
        else:
            total_to_process = sum([
                1 + (1 if state["supplier_settings"][key]["credit"] else 0)
                for key in selected_suppliers
            ])

            st.success(f"✅ Klaar om {len(selected_suppliers)} leverancier(s) te verwerken ({total_to_process} bestanden totaal)")

            if st.button("🚀 START VERWERKING", type="primary", key=f"{customer_name}_start"):
                state["results"] = []

                progress_bar = st.progress(0)
                status_text = st.empty()

                processed = 0

                for supplier_key in selected_suppliers:
                    status_text.text(f"Verwerken: {SUPPLIERS[supplier_key]['naam']} (debet)...")

                    result_debet = process_supplier(
                        state["bronbestand"],
                        state["templates"][supplier_key],
                        supplier_key,
                        state["supplier_settings"][supplier_key]["periode"],
                        False,
                        state["template_filenames"].get(supplier_key),
                        is_credit_sheet=False
                    )

                    state["results"].append(result_debet)
                    processed += 1
                    progress_bar.progress(processed / total_to_process)

                    if state["supplier_settings"][supplier_key]["credit"]:
                        status_text.text(f"Verwerken: {SUPPLIERS[supplier_key]['naam']} (credit)...")

                        result_credit = process_supplier(
                            state["bronbestand"],
                            state["templates"][supplier_key],
                            supplier_key,
                            state["supplier_settings"][supplier_key]["periode"],
                            True,
                            state["template_filenames"].get(supplier_key),
                            is_credit_sheet=True,
                            credit_template_bytes=state["credit_templates"].get(supplier_key),
                            credit_template_filename=state["credit_template_filenames"].get(supplier_key)
                        )

                        state["results"].append(result_credit)
                        processed += 1
                        progress_bar.progress(processed / total_to_process)

                status_text.text("✅ Verwerking voltooid")
                st.balloons()
                st.rerun()

# =====================================================
# UI: Euromaster SEFE tab 
# =====================================================
def render_euromaster_sefe():
    tab1, tab2 = st.tabs(["📁 Upload", "📊 Validatie"])

    state_key = "euromaster_sefe_state"
    if state_key not in st.session_state:
        st.session_state[state_key] = {
            "report": None,
            "excel_bytes": None,
            "gewijzigd": False
        }

    eu_state = st.session_state[state_key]

    with tab1:
        st.markdown("## 📁 Euromaster SEFE upload")
        uploaded = st.file_uploader(
            "Upload csv bestand",
            type=["csv"],
            key="euromaster_csv_upload"
        )


        if uploaded:
            excel_bytes, report, gewijzigd = preprocess_euromaster_sefe(uploaded.read())
            eu_state["report"] = report
            eu_state["excel_bytes"] = excel_bytes
            eu_state["gewijzigd"] = gewijzigd

            if report["fouten"]:
                st.error("❌ Fouten gevonden. Zie Validatie tab.")
            else:
                if gewijzigd:
                    st.success("✅ Validatie gelukt. Wijzigingen toegepast. Je kunt downloaden.")
                else:
                    st.info("Geen wijzigingen nodig. Bestand kan zonder opslaan worden afgesloten.")

    with tab2:
        st.markdown("## 📊 Validatie-overzicht")
        if not eu_state["report"]:
            st.info("Upload eerst een zipbestand in de Upload tab.")
            return

        report = eu_state["report"]

        if report["correcties"]:
            st.success("Correcties")
            for c in report["correcties"]:
                st.write("•", c)

        if report["waarschuwingen"]:
            st.warning("Waarschuwingen")
            for w in report["waarschuwingen"]:
                st.write("•", w)

        if report["fouten"]:
            st.error("Fouten")
            for f in report["fouten"]:
                st.write("•", f)

        if (not report["fouten"]) and eu_state["gewijzigd"] and eu_state["excel_bytes"]:
            st.download_button(
                "Download SEFE Excel",
                data=eu_state["excel_bytes"],
                file_name="Euromaster_SeFe.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="euromaster_download"
            )

# =====================================================
# MAIN
# =====================================================
def main():
    st.markdown("""
        <div class="main-header">
            <h1>⚡ Betaalbestanden Generator</h1>
            <p>Energiemissie</p>
        </div>
    """, unsafe_allow_html=True)

    # Klant tabs
    tab_pnh, tab_ggz, tab_euro = st.tabs(["Provincie Noord-Holland", "GGZ Centraal", "Euromaster"])

    with tab_pnh:
        st.markdown("## Provincie Noord-Holland")
        render_standard_customer_flow("Provincie Noord-Holland", CUSTOMER_SUPPLIERS["Provincie Noord-Holland"])

    with tab_ggz:
        st.markdown("## GGZ Centraal")
        render_standard_customer_flow("GGZ Centraal", CUSTOMER_SUPPLIERS["GGZ Centraal"])

    with tab_euro:
        st.markdown("## Euromaster SEFE")
        render_euromaster_sefe()

if __name__ == "__main__":
    main()