"""
Verzamelblad Generator
"""
import streamlit as st
import pandas as pd
import os
import tempfile
import datetime as dt
from openpyxl import load_workbook
import re
from io import BytesIO
import random

# Pagina configuratie
st.set_page_config(
    page_title="Verzamelblad Generator - Energiemissie",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS voor mooie styling
st.markdown("""
<style>
    /* Hoofdkleuren */
    :root {
        --primary-color: #667eea;
        --secondary-color: #764ba2;
        --success-color: #4ade80;
        --error-color: #ef4444;
    }
    
    /* Header styling */
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 1rem;
        color: white;
        margin-bottom: 2rem;
        text-align: center;
    }
    
    .main-header h1 {
        color: white !important;
        font-size: 3rem;
        margin: 0;
        font-weight: 800;
    }
    
    .main-header p {
        color: rgba(255, 255, 255, 0.9);
        font-size: 1.2rem;
        margin-top: 0.5rem;
    }
    
    /* Card styling */
    .stCard {
        background: white;
        border-radius: 1rem;
        padding: 1.5rem;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        border: 1px solid #e5e7eb;
    }
    
    /* Supplier cards */
    .supplier-card {
        background: white;
        border-left: 4px solid #667eea;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
    }
    
    .supplier-card.selected {
        border-left-color: #4ade80;
        background: #f0fdf4;
    }
    
    /* Success/Error messages */
    .success-box {
        background: #dcfce7;
        border-left: 4px solid #16a34a;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    
    .error-box {
        background: #fee2e2;
        border-left: 4px solid #dc2626;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    
    .warning-box {
        background: #fef3c7;
        border-left: 4px solid #f59e0b;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    
    .info-box {
        background: #dbeafe;
        border-left: 4px solid #3b82f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
    }
    
    /* Metrics styling */
    .metric-card {
        background: #f9fafb;
        padding: 1rem;
        border-radius: 0.5rem;
        text-align: center;
    }
    
    .metric-label {
        color: #6b7280;
        font-size: 0.875rem;
        margin-bottom: 0.25rem;
    }
    
    .metric-value {
        color: #1f2937;
        font-size: 1.5rem;
        font-weight: 700;
    }
    
    /* Buttons */
    .stButton>button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 0.5rem;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 1rem;
        width: 100%;
        transition: all 0.3s ease;
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 10px 20px rgba(102, 126, 234, 0.3);
    }
    
    /* File uploader */
    .uploadedFile {
        background: #f0fdf4;
        border: 2px solid #4ade80;
        border-radius: 0.5rem;
        padding: 0.5rem;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: #f9fafb;
        border-radius: 0.5rem;
        font-weight: 600;
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 2rem;
    }
    
    .stTabs [data-baseweb="tab"] {
        font-weight: 600;
        font-size: 1rem;
    }
</style>
""", unsafe_allow_html=True)

# LEVERANCIERS 
SUPPLIERS = {
    "kenter": {
        "naam": "Kenter",
        "tabnaam": "NL59INGB0006779355_D",
        "tabnaam_credit": "NL59INGB0006779355_C",
        "color": "#FF6B6B"
    },
    "liander": {
        "naam": "Liander",
        "tabnaam": "NL95INGB0000005585_D",
        "tabnaam_credit": "NL95INGB0000005585_C",
        "color": "#4ECDC4"
    },
    "vattenfall": {
        "naam": "Vattenfall",
        "tabnaam": "NL42INGB0000827935_D",
        "tabnaam_credit": "NL14ABNA0242263240_C",
        "color": "#FFE66D"
    },
    "eneco": {
        "naam": "Eneco",
        "tabnaam": "NL13ABNA0640000797_D",
        "tabnaam_credit": "NL13ABNA0640000797_C",
        "color": "#95E1D3"
    },
    "vitens": {
        "naam": "Vitens",
        "tabnaam": "NL94INGB0000869000_D",
        "tabnaam_credit": "NL94INGB0000869000_C",
        "color": "#A8E6CF"
    }
}


# =====================================================================
# FUNCTIE OM RANDOM FACTUUR TE KRIJGEN
# =====================================================================
def get_random_invoice(bronbestand_bytes, tabnaam):
    """Haal een random factuurnummer op uit het bronbestand voor referentie"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(bronbestand_bytes)
            tmp_path = tmp.name
        
        try:
            df_raw = pd.read_excel(tmp_path, sheet_name=tabnaam, header=None).dropna(how="all")
            df_headers = df_raw.iloc[0]
            df_data = df_raw[1:].copy()
            df_data.columns = df_headers
            
            # Kolom C = factuurnummers (index 2)
            if len(df_data) > 0:
                factuurnummers = df_data.iloc[:, 2].astype(str).tolist()
                if factuurnummers:
                    return random.choice(factuurnummers)
            return None
        finally:
            try:
                os.unlink(tmp_path)
            except:
                pass
    except:
        return None

def process_supplier(
    bronbestand_bytes,
    template_bytes,
    supplier_key: str,
    periode: str,
    credit: bool,
    template_filename: str = None,
    is_credit_sheet: bool = False,
    credit_template_bytes = None,
    credit_template_filename: str = None
):
    """Verwerk een enkele leverancier (debet of credit)"""
    
    # Kies het juiste tabblad
    if is_credit_sheet:
        tabnaam = SUPPLIERS[supplier_key]["tabnaam_credit"]
        template_to_use = credit_template_bytes if credit_template_bytes else template_bytes
        filename_to_use = credit_template_filename if credit_template_filename else template_filename
    else:
        tabnaam = SUPPLIERS[supplier_key]["tabnaam"]
        template_to_use = template_bytes
        filename_to_use = template_filename
    
    # Datum in ddmmyy formaat (6 cijfers)
    datum_excel = dt.date.today().strftime("%d%m%y")
    datum_tekst = dt.date.today().strftime("%d-%m-%Y")
    
    bron_path = None
    tpl_path = None
    
    try:
        # Maak tijdelijke bestanden
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as bron_tmp:
            bron_tmp.write(bronbestand_bytes)
            bron_path = bron_tmp.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tpl_tmp:
            tpl_tmp.write(template_to_use)
            tpl_path = tpl_tmp.name
        
        # Check tabblad
        xls_file = pd.ExcelFile(bron_path)
        sheets = xls_file.sheet_names
        xls_file.close()
        
        if tabnaam not in sheets:
            # Cleanup
            try:
                if bron_path and os.path.exists(bron_path):
                    os.unlink(bron_path)
                if tpl_path and os.path.exists(tpl_path):
                    os.unlink(tpl_path)
            except:
                pass
            
            return {
                "success": False,
                "message": f"Tabblad {tabnaam} niet gevonden",
                "supplier": SUPPLIERS[supplier_key]["naam"],
                "is_credit": is_credit_sheet
            }
        
        # Lees data
        df_raw = pd.read_excel(bron_path, sheet_name=tabnaam, header=None).dropna(how="all")
        df_headers = df_raw.iloc[0]
        df_data = df_raw[1:].copy()
        df_data.columns = df_headers
        
        # Check of er data is
        if len(df_data) == 0:
            try:
                if bron_path and os.path.exists(bron_path):
                    os.unlink(bron_path)
                if tpl_path and os.path.exists(tpl_path):
                    os.unlink(tpl_path)
            except:
                pass
            
            return {
                "success": False,
                "message": "Geen data gevonden in tabblad",
                "supplier": SUPPLIERS[supplier_key]["naam"],
                "is_credit": is_credit_sheet,
                "no_data": True
            }
        
        # Open template
        wb = load_workbook(tpl_path)
        ws_spec = wb["Specificatie"]
        
        # Specificatie leegmaken
        last = 1
        while ws_spec.cell(last + 1, 1).value not in (None, ""):
            last += 1
        if last > 1:
            ws_spec.delete_rows(2, last - 1)
        
        # Nieuwe data
        for r, row in enumerate(df_data.values, start=2):
            for c, v in enumerate(row, start=1):
                ws_spec.cell(r, c, v)
        
        # Verzamelblad bijwerken
        ws_verz = wb["Verzamelblad"]
        ws_verz["B4"].value = datum_tekst
        
        if supplier_key in ("kenter", "liander", "vattenfall"):
            ws_verz["C24"].value = f"{supplier_key.capitalize()}_{datum_excel}"
        else:
            old = ws_verz["C24"].value
            if isinstance(old, str):
                new = re.sub(r"\d{6,8}", datum_excel, old, 1)
                ws_verz["C24"].value = new
            else:
                ws_verz["C24"].value = f"{supplier_key.upper()}_{datum_excel}"
        
        ws_verz["C26"].value = periode
        
        # Bepaal output bestandsnaam
        if filename_to_use:
            template_basename = filename_to_use.replace('.xlsx', '')
        else:
            template_basename = SUPPLIERS[supplier_key]['naam']
            if is_credit_sheet:
                template_basename += "_Credit"
        
        # Vervang laatste datum
        pattern = r'_(\d{6,8})(?!.*_\d)'
        
        if re.search(pattern, template_basename):
            output_filename_base = re.sub(pattern, f'_{datum_excel}', template_basename)
        else:
            output_filename_base = f"{template_basename}_{datum_excel}"
        
        output_filename = f"{output_filename_base}.xlsx"
        
        # Opslaan naar bytes
        output = BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)
        excel_bytes = output.read()
        
        # Somcontrole
        wb_check = load_workbook(BytesIO(excel_bytes))
        ws_new = wb_check["Specificatie"]
        
        rows_new = []
        r = 2
        while ws_new.cell(r, 1).value not in (None, ""):
            row = [ws_new.cell(r, c).value for c in range(1, df_data.shape[1] + 1)]
            rows_new.append(row)
            r += 1
        
        wb_check.close()
        df_spec_new = pd.DataFrame(rows_new, columns=df_data.columns)
        
        # Totalen
        kol_excl = df_data.columns.get_loc("Excl. BTW")
        kol_btw = df_data.columns.get_loc("BTW")
        kol_incl = df_data.columns.get_loc("Incl. BTW")
        
        sum_excl_bron = df_data.iloc[:, kol_excl].astype(float).sum()
        sum_btw_bron = df_data.iloc[:, kol_btw].astype(float).sum()
        sum_incl_bron = df_data.iloc[:, kol_incl].astype(float).sum()
        
        sum_excl_new = df_spec_new.iloc[:, kol_excl].astype(float).sum()
        sum_btw_new = df_spec_new.iloc[:, kol_btw].astype(float).sum()
        sum_incl_new = df_spec_new.iloc[:, kol_incl].astype(float).sum()
        
        # Check
        bedragen_kloppen = (
            abs(sum_excl_bron - sum_excl_new) < 0.01 and
            abs(sum_btw_bron - sum_btw_new) < 0.01 and
            abs(sum_incl_bron - sum_incl_new) < 0.01
        )
        
        # Cleanup
        try:
            if bron_path and os.path.exists(bron_path):
                os.unlink(bron_path)
        except Exception:
            pass
        
        try:
            if tpl_path and os.path.exists(tpl_path):
                os.unlink(tpl_path)
        except Exception:
            pass
        
        # PDF generatie
        pdf_bytes = None
        pdf_filename = None
        
        if bedragen_kloppen:
            try:
                import win32com.client as win32
                
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
                    tmp_excel.write(excel_bytes)
                    tmp_excel_path = tmp_excel.name
                
                pdf_temp_path = tmp_excel_path.replace('.xlsx', '.pdf')
                
                try:
                    excel_app = win32.Dispatch("Excel.Application")
                    excel_app.Visible = False
                    excel_app.DisplayAlerts = False
                    
                    wb_pdf = excel_app.Workbooks.Open(tmp_excel_path, ReadOnly=True)
                    ws_pdf = wb_pdf.Worksheets("Verzamelblad")
                    ws_pdf.ExportAsFixedFormat(0, pdf_temp_path)
                    
                    wb_pdf.Close(False)
                    excel_app.Quit()
                    
                    with open(pdf_temp_path, 'rb') as f:
                        pdf_bytes = f.read()
                    
                    pdf_filename = output_filename.replace('.xlsx', '.pdf')
                    
                    try:
                        os.unlink(tmp_excel_path)
                        os.unlink(pdf_temp_path)
                    except:
                        pass
                    
                except Exception:
                    try:
                        if 'wb_pdf' in locals():
                            wb_pdf.Close(False)
                        if 'excel_app' in locals():
                            excel_app.Quit()
                        if os.path.exists(tmp_excel_path):
                            os.unlink(tmp_excel_path)
                        if os.path.exists(pdf_temp_path):
                            os.unlink(pdf_temp_path)
                    except:
                        pass
                    
            except ImportError:
                pass
            except Exception:
                pass
        
        return {
            "success": bedragen_kloppen,
            "message": "Bedragen kloppen ✓" if bedragen_kloppen else "Bedragen kloppen NIET ✗",
            "supplier": SUPPLIERS[supplier_key]["naam"],
            "excl": sum_excl_new,
            "btw": sum_btw_new,
            "incl": sum_incl_new,
            "excl_bron": sum_excl_bron,
            "btw_bron": sum_btw_bron,
            "incl_bron": sum_incl_bron,
            "excel_bytes": excel_bytes,
            "filename": output_filename,
            "pdf_bytes": pdf_bytes,
            "pdf_filename": pdf_filename,
            "is_credit": is_credit_sheet
        }
        
    except Exception as e:
        try:
            if bron_path and os.path.exists(bron_path):
                os.unlink(bron_path)
        except:
            pass
        
        try:
            if tpl_path and os.path.exists(tpl_path):
                os.unlink(tpl_path)
        except:
            pass
        
        return {
            "success": False,
            "message": f"Fout: {str(e)}",
            "supplier": SUPPLIERS[supplier_key]["naam"],
            "is_credit": is_credit_sheet
        }


def main():
    # Header
    st.markdown("""
        <div class="main-header">
            <h1>⚡ Verzamelblad Generator</h1>
            <p>Energiemissie - Automatische verwerking voor iedereen</p>
        </div>
    """, unsafe_allow_html=True)
    
    # Sidebar met info
    with st.sidebar:
        st.markdown("---")
        st.markdown("### 📋 Instructies")
        st.markdown("""
        1. **Upload bronbestand** (DB Energie export)
        2. **Selecteer klant**
        3. **Upload templates** voor leveranciers
        4. **Upload credit templates** (optioneel)
        5. **Configureer periode** per leverancier
        6. **Klik op verwerken**
        7. **Download resultaten**
        """)
        st.markdown("---")
        st.markdown("### ℹ️ Info")
        st.info("Deze app verwerkt automatisch verzamelbladen (debet en credit) en controleert de bedragen.")
        
    # Tabs voor workflow
    tab1, tab2, tab3, tab4 = st.tabs(["📁 Upload", "🏢 Klant", "⚙️ Configuratie", "🚀 Verwerken"])
    
    # Initialize session state
    if 'bronbestand' not in st.session_state:
        st.session_state.bronbestand = None
    if 'klant' not in st.session_state:
        st.session_state.klant = None
    if 'templates' not in st.session_state:
        st.session_state.templates = {}
    if 'template_filenames' not in st.session_state:
        st.session_state.template_filenames = {}
    if 'credit_templates' not in st.session_state:
        st.session_state.credit_templates = {}
    if 'credit_template_filenames' not in st.session_state:
        st.session_state.credit_template_filenames = {}
    if 'supplier_settings' not in st.session_state:
        st.session_state.supplier_settings = {
            key: {"selected": False, "credit": False, "periode": ""}
            for key in SUPPLIERS.keys()
        }
    if 'results' not in st.session_state:
        st.session_state.results = []
    
    # TAB 1: Upload bronbestand
    with tab1:
        st.markdown("## 📁 Bronbestand uploaden")
        st.markdown("Upload het exportbestand vanuit DB Energie (.xlsx)")
        
        uploaded_bron = st.file_uploader(
            "Kies bronbestand",
            type=["xlsx", "xls"],
            key="bron_upload",
            help="Dit is het bestand dat je exporteert uit DB Energie"
        )
        
        if uploaded_bron:
            st.session_state.bronbestand = uploaded_bron.read()
            
            st.success(f"✅ Bestand geladen: **{uploaded_bron.name}**")
            
            # Check welke leveranciers beschikbaar zijn
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    tmp.write(st.session_state.bronbestand)
                    tmp_path = tmp.name
                
                xls = pd.ExcelFile(tmp_path)
                sheets = set(xls.sheet_names)
                xls.close()
                
                st.markdown("### 🔍 Gevonden leveranciers:")
                
                # Maak twee rijen: debet en credit
                st.markdown("**📊 Debet facturen:**")
                cols_debet = st.columns(len(SUPPLIERS))
                
                for idx, (key, supplier) in enumerate(SUPPLIERS.items()):
                    with cols_debet[idx]:
                        if supplier["tabnaam"] in sheets:
                            st.markdown(f"✅ **{supplier['naam']}**")
                            st.session_state.supplier_settings[key]["selected"] = True
                        else:
                            st.markdown(f"❌ {supplier['naam']}")
                            st.session_state.supplier_settings[key]["selected"] = False
                
                st.markdown("**💳 Credit facturen:**")
                cols_credit = st.columns(len(SUPPLIERS))
                
                has_credits = False
                for idx, (key, supplier) in enumerate(SUPPLIERS.items()):
                    with cols_credit[idx]:
                        if supplier["tabnaam_credit"] in sheets:
                            st.markdown(f"✅ **{supplier['naam']}**")
                            st.session_state.supplier_settings[key]["credit"] = True
                            has_credits = True
                        else:
                            st.markdown(f"⚪ {supplier['naam']}")
                
                if has_credits:
                    st.info("💡 Credit facturen gevonden! Upload credit templates in de Configuratie tab.")
                
            except Exception as e:
                st.error(f"Fout bij het lezen van het bestand: {e}")
            finally:
                if tmp_path:
                    try:
                        os.unlink(tmp_path)
                    except Exception:
                        pass
        else:
            st.info("👆 Upload eerst een bronbestand om te beginnen")
    
    # TAB 2: Klant selectie
    with tab2:
        st.markdown("## 🏢 Selecteer klant")
        
        if not st.session_state.bronbestand:
            st.warning("⚠️ Upload eerst een bronbestand in de 'Upload' tab")
        else:
            klanten = [
                "Provincie Noord-Holland",
                "GGZ Centraal",
            ]
            
            selected_klant = st.selectbox(
                "Kies de klant",
                ["Selecteer een klant..."] + klanten,
                index=0
            )
            
            if selected_klant != "Selecteer een klant...":
                st.session_state.klant = selected_klant
                st.success(f"✅ Klant geselecteerd: **{selected_klant}**")
    
    # TAB 3: Configuratie
    with tab3:
        st.markdown("## ⚙️ Configureer leveranciers")
        
        if not st.session_state.bronbestand or not st.session_state.klant:
            st.warning("⚠️ Voltooi eerst de vorige stappen")
        else:
            for supplier_key, supplier in SUPPLIERS.items():
                with st.expander(
                    f"🔌 {supplier['naam']}" + 
                    (" ✅ Actief" if st.session_state.supplier_settings[supplier_key]["selected"] else " ⏸️ Inactief"),
                    expanded=st.session_state.supplier_settings[supplier_key]["selected"]
                ):
                    col1, col2 = st.columns([1, 3])
                    
                    with col1:
                        selected = st.checkbox(
                            "Verwerken",
                            value=st.session_state.supplier_settings[supplier_key]["selected"],
                            key=f"sel_{supplier_key}"
                        )
                        st.session_state.supplier_settings[supplier_key]["selected"] = selected
                    
                    if selected:
                        with col2:
                            has_credit = st.session_state.supplier_settings[supplier_key]["credit"]
                            if has_credit:
                                st.success(f"💳 Credit facturen gevonden! ({supplier['tabnaam_credit']})")
                        
                        # Toon random factuur voor periode opzoeken
                        st.markdown("---")
                        st.markdown("**🔍 Referentie factuurnummer:**")
                        
                        # Debet factuur
                        random_invoice_debet = get_random_invoice(st.session_state.bronbestand, supplier["tabnaam"])
                        if random_invoice_debet:
                            st.markdown(f"""
                            <div class="info-box">
                                <strong>📊 Debet:</strong> {random_invoice_debet}<br>
                                <em style="font-size: 0.9em; color: #6b7280;">Gebruik dit factuurnummer om de periode op te zoeken in DB EFactuur</em>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        # Credit factuur (als beschikbaar)
                        if has_credit:
                            random_invoice_credit = get_random_invoice(st.session_state.bronbestand, supplier["tabnaam_credit"])
                            if random_invoice_credit:
                                st.markdown(f"""
                                <div class="info-box">
                                    <strong>💳 Credit:</strong> {random_invoice_credit}<br>
                                    <em style="font-size: 0.9em; color: #6b7280;">Gebruik dit factuurnummer om de periode op te zoeken in DB EFactuur</em>
                                </div>
                                """, unsafe_allow_html=True)
                        
                        st.markdown("---")
                        
                        periode = st.text_input(
                            "Periode (bijv. 01-12-2025 t/m 31-12-2025)",
                            value=st.session_state.supplier_settings[supplier_key]["periode"],
                            key=f"periode_{supplier_key}",
                            help="Zoek het factuurnummer hierboven op in DB EFactuur om de juiste periode te vinden"
                        )
                        st.session_state.supplier_settings[supplier_key]["periode"] = periode
                        
                        st.markdown("**📊 Debet Template:**")
                        template = st.file_uploader(
                            f"Template voor {supplier['naam']} (debet)",
                            type=["xlsx"],
                            key=f"template_{supplier_key}"
                        )
                        
                        if template:
                            st.session_state.templates[supplier_key] = template.read()
                            st.session_state.template_filenames[supplier_key] = template.name
                            st.success(f"✅ Debet template geladen: {template.name}")
                        elif supplier_key in st.session_state.templates:
                            st.info(f"✅ Debet template al geladen: {st.session_state.template_filenames.get(supplier_key, 'onbekend')}")
                        
                        # Credit template upload (alleen als credit beschikbaar)
                        if has_credit:
                            st.markdown("**💳 Credit Template:**")
                            credit_template = st.file_uploader(
                                f"Credit template voor {supplier['naam']}",
                                type=["xlsx"],
                                key=f"credit_template_{supplier_key}"
                            )
                            
                            if credit_template:
                                st.session_state.credit_templates[supplier_key] = credit_template.read()
                                st.session_state.credit_template_filenames[supplier_key] = credit_template.name
                                st.success(f"✅ Credit template geladen: {credit_template.name}")
                            elif supplier_key in st.session_state.credit_templates:
                                st.info(f"✅ Credit template al geladen: {st.session_state.credit_template_filenames.get(supplier_key, 'onbekend')}")
    
    # TAB 4: Verwerken
    with tab4:
        st.markdown("## 🚀 Bestanden verwerken")
        
        # Check of alles klaar is
        selected_suppliers = [
            key for key, settings in st.session_state.supplier_settings.items()
            if settings["selected"]
        ]
        
        missing_templates = [
            SUPPLIERS[key]["naam"] for key in selected_suppliers
            if key not in st.session_state.templates
        ]
        
        missing_credit_templates = [
            SUPPLIERS[key]["naam"] for key in selected_suppliers
            if st.session_state.supplier_settings[key]["credit"] and key not in st.session_state.credit_templates
        ]
        
        can_process = (
            st.session_state.bronbestand and
            st.session_state.klant and
            len(selected_suppliers) > 0 and
            len(missing_templates) == 0 and
            len(missing_credit_templates) == 0
        )
        
        if not can_process:
            st.warning("⚠️ Vul eerst alle vereiste velden in:")
            issues = []
            if not st.session_state.bronbestand:
                issues.append("- Upload bronbestand")
            if not st.session_state.klant:
                issues.append("- Selecteer klant")
            if len(selected_suppliers) == 0:
                issues.append("- Selecteer minimaal 1 leverancier")
            if missing_templates:
                issues.append(f"- Upload debet templates voor: {', '.join(missing_templates)}")
            if missing_credit_templates:
                issues.append(f"- Upload credit templates voor: {', '.join(missing_credit_templates)}")
            
            st.markdown("\n".join(issues))
        else:
            # Toon overzicht
            total_to_process = sum([
                1 + (1 if st.session_state.supplier_settings[key]["credit"] else 0)
                for key in selected_suppliers
            ])
            
            st.success(f"✅ Klaar om {len(selected_suppliers)} leverancier(s) te verwerken ({total_to_process} bestanden totaal)!")
            
            if st.button("🚀 START VERWERKING", type="primary"):
                st.session_state.results = []
                
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                processed = 0
                
                for supplier_key in selected_suppliers:
                    # Verwerk debet
                    status_text.text(f"Verwerken: {SUPPLIERS[supplier_key]['naam']} (debet)...")
                    
                    result_debet = process_supplier(
                        st.session_state.bronbestand,
                        st.session_state.templates[supplier_key],
                        supplier_key,
                        st.session_state.supplier_settings[supplier_key]["periode"],
                        False,
                        st.session_state.template_filenames.get(supplier_key),
                        is_credit_sheet=False
                    )
                    
                    st.session_state.results.append(result_debet)
                    processed += 1
                    progress_bar.progress(processed / total_to_process)
                    
                    # Verwerk credit (als beschikbaar)
                    if st.session_state.supplier_settings[supplier_key]["credit"]:
                        status_text.text(f"Verwerken: {SUPPLIERS[supplier_key]['naam']} (credit)...")
                        
                        result_credit = process_supplier(
                            st.session_state.bronbestand,
                            st.session_state.templates[supplier_key],
                            supplier_key,
                            st.session_state.supplier_settings[supplier_key]["periode"],
                            True,
                            st.session_state.template_filenames.get(supplier_key),
                            is_credit_sheet=True,
                            credit_template_bytes=st.session_state.credit_templates.get(supplier_key),
                            credit_template_filename=st.session_state.credit_template_filenames.get(supplier_key)
                        )
                        
                        st.session_state.results.append(result_credit)
                        processed += 1
                        progress_bar.progress(processed / total_to_process)
                
                status_text.text("✅ Verwerking voltooid!")
                st.balloons()
        
        # Toon resultaten
        if st.session_state.results:
            st.markdown("---")
            st.markdown("## 📊 Resultaten & Downloads")
            
            for result in st.session_state.results:
                # Label voor credit/debet
                type_label = "💳 CREDIT" if result.get("is_credit") else "📊 DEBET"
                
                # Skip resultaten zonder data
                if result.get("no_data"):
                    continue
                
                if result["success"]:
                    with st.container():
                        st.markdown(f"""
                        <div class="success-box">
                            <h3 style="margin:0;">✅ {result['supplier']} {type_label}</h3>
                            <p style="margin:0.5rem 0 0 0;">{result['message']}</p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        col1, col2, col3 = st.columns(3)
                        
                        with col1:
                            st.metric("Excl. BTW", f"€ {result['excl']:,.2f}")
                        with col2:
                            st.metric("BTW", f"€ {result['btw']:,.2f}")
                        with col3:
                            st.metric("Incl. BTW", f"€ {result['incl']:,.2f}")
                        
                        # Download buttons
                        st.markdown("**📥 Downloads:**")
                        col_excel, col_pdf = st.columns(2)
                        
                        with col_excel:
                            st.download_button(
                                f"📊 {result['filename']}",
                                data=result["excel_bytes"],
                                file_name=result["filename"],
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"excel_{result['supplier']}_{result.get('is_credit', False)}"
                            )
                        
                        with col_pdf:
                            if result.get("pdf_bytes"):
                                st.download_button(
                                    f"📄 {result['pdf_filename']}",
                                    data=result["pdf_bytes"],
                                    file_name=result["pdf_filename"],
                                    mime="application/pdf",
                                    use_container_width=True,
                                    key=f"pdf_{result['supplier']}_{result.get('is_credit', False)}"
                                )
                            else:
                                st.info("📄 PDF: Excel niet beschikbaar", icon="ℹ️")
                        
                        st.markdown("---")
                        
                else:
                    with st.container():
                        st.markdown(f"""
                        <div class="error-box">
                            <h3 style="margin:0;">❌ {result['supplier']} {type_label}</h3>
                            <p style="margin:0.5rem 0 0 0;">{result['message']}</p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if "excl" in result:
                            st.markdown("**Vergelijking bedragen:**")
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.markdown("**Bronbestand:**")
                                st.text(f"Excl: € {result['excl_bron']:,.2f}")
                                st.text(f"BTW:  € {result['btw_bron']:,.2f}")
                                st.text(f"Incl: € {result['incl_bron']:,.2f}")
                            
                            with col2:
                                st.markdown("**Nieuw:**")
                                st.text(f"Excl: € {result['excl']:,.2f}")
                                st.text(f"BTW:  € {result['btw']:,.2f}")
                                st.text(f"Incl: € {result['incl']:,.2f}")
                            
                            with col3:
                                st.markdown("**Verschil:**")
                                st.text(f"Excl: € {result['excl']-result['excl_bron']:,.2f}")
                                st.text(f"BTW:  € {result['btw']-result['btw_bron']:,.2f}")
                                st.text(f"Incl: € {result['incl']-result['incl_bron']:,.2f}")
                        
                        st.markdown("---")


if __name__ == "__main__":
    main()

